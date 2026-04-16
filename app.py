import streamlit as st
import pandas as pd
import io
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(
    page_title="Indeed請求明細ジェネレーター",
    page_icon="📊",
    layout="wide"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;700&family=Inter:wght@400;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Noto Sans JP', sans-serif; }
.title-bar { background: linear-gradient(135deg, #1F4E79, #2E75B6); padding: 20px 28px 16px 28px; border-radius: 12px; margin-bottom: 10px; }
.main-title { font-family: 'Inter', sans-serif; font-size: 1.8rem; font-weight: 700; color: #FFFFFF; margin: 0; padding: 0; }
.sub-title { color: #CBD5E1; font-size: 0.88rem; margin-top: 4px; }
.section-card { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 12px 18px; margin-bottom: 10px; }
.stButton > button { background-color: #1F4E79; color: white; border: none; border-radius: 8px; font-size: 1rem; font-weight: 700; padding: 10px 28px; width: 100%; transition: all 0.2s; }
.stButton > button:hover { background-color: #2E75B6; transform: translateY(-1px); box-shadow: 0 4px 12px rgba(31,78,121,0.3); }
.result-box { background: #e8f5e9; border: 1px solid #a5d6a7; border-radius: 8px; padding: 14px; color: #2e7d32; font-weight: 600; margin-top: 10px; }
.error-box { background: #ffebee; border: 1px solid #ef9a9a; border-radius: 8px; padding: 14px; color: #c62828; margin-top: 10px; }
hr { margin: 8px 0 !important; border-color: #e2e8f0 !important; }
</style>
""", unsafe_allow_html=True)

def get_drive_service():
    creds_info = dict(st.secrets["GOOGLE_SERVICE_ACCOUNT"])
    if "private_key" in creds_info:
        creds_info["private_key"] = creds_info["private_key"].replace("\\n", "\n")
    creds = service_account.Credentials.from_service_account_info(
        creds_info, scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    return build("drive", "v3", credentials=creds)

def list_files_in_folder(service, folder_id):
    results = service.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id, name, mimeType)", orderBy="name"
    ).execute()
    return results.get("files", [])

def download_file(service, file_id):
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf

def parse_yen(val):
    if pd.isna(val): return 0
    s = str(val).replace('￥','').replace(',','').strip()
    try: return int(float(s))
    except: return 0

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style='medium', color='1F4E79')
    return Border(left=s, right=s, top=s, bottom=s)

def create_billing_excel(client_name, inv_df, csv_df, month_label):
    hits_inv = inv_df[inv_df['Client name'].str.contains(client_name, na=False)].copy()
    if len(hits_inv) == 0:
        return None, f"「{client_name}」に一致するクライアントが見つかりませんでした"
    target_ids = hits_inv['Employer ID'].tolist()
    if 'メジャー ネーム' in csv_df.columns:
        cost_df = csv_df[(csv_df['アカウントID'].isin(target_ids)) & (csv_df['メジャー ネーム'] == '合計費用')].copy()
        cost_df['合計費用_数値'] = cost_df['メジャー バリュー'].fillna(0).astype(float).astype(int)
    elif '合計費用' in csv_df.columns:
        cost_df = csv_df[csv_df['アカウントID'].isin(target_ids)].copy()
        cost_df['合計費用_数値'] = cost_df['合計費用'].apply(parse_yen)
    else:
        return None, "CSVの形式を認識できませんでした"
    merged = hits_inv.merge(
        cost_df[['アカウントID','アカウント名','キャンペーン名','キャンペーン開始日','キャンペーン終了日 (指定した日付)','キャンペーンステータス','合計費用_数値']],
        left_on='Employer ID', right_on='アカウントID', how='left'
    )
    diff = hits_inv['費消額'].sum() - cost_df['合計費用_数値'].sum()
    HEADER_BG, WHITE, GRAY, TOTAL_BG = '1F4E79', 'FFFFFF', 'F5F5F5', 'FFF2CC'
    wb = Workbook()
    ws = wb.active
    ws.title = '請求明細'
    ws.merge_cells('A1:G1')
    ws['A1'] = f'{client_name}　Indeed請求明細　{month_label}'
    ws['A1'].font = Font(name='Arial', size=13, bold=True, color=WHITE)
    ws['A1'].fill = PatternFill('solid', fgColor=HEADER_BG)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    headers = ['アカウント名','キャンペーン名','開始日','終了日','ステータス','キャンペーン費消額（円）','アカウント合計費消額（円）']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name='Arial', size=10, bold=True, color=WHITE)
        cell.fill = PatternFill('solid', fgColor='2E75B6')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[2].height = 28
    account_order = merged['Employer ID'].unique()
    fill_white = PatternFill('solid', fgColor=WHITE)
    fill_gray = PatternFill('solid', fgColor=GRAY)
    row = 3
    for i, acc_id in enumerate(account_order):
        grp = merged[merged['Employer ID'] == acc_id].reset_index(drop=True)
        acc_name = grp['Client name'].iloc[0]
        fill = fill_gray if i % 2 == 0 else fill_white
        start_row = row
        for j, r in grp.iterrows():
            camp_fee = r['合計費用_数値'] if pd.notna(r['合計費用_数値']) else 0
            ws.cell(row=row, column=1, value=acc_name if j == 0 else '')
            ws.cell(row=row, column=2, value=r.get('キャンペーン名',''))
            ws.cell(row=row, column=3, value=r.get('キャンペーン開始日',''))
            ws.cell(row=row, column=4, value=r.get('キャンペーン終了日 (指定した日付)',''))
            ws.cell(row=row, column=5, value=r.get('キャンペーンステータス',''))
            ws.cell(row=row, column=6, value=int(camp_fee))
            ws.cell(row=row, column=7, value='')
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name='Arial', size=9)
                cell.fill = fill
                cell.border = thin_border()
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if col in (6, 7):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                if col in (3, 4):
                    cell.number_format = 'YYYY/MM/DD'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 22
            row += 1
        end_row = row - 1
        if start_row < end_row:
            ws.merge_cells(f'A{start_row}:A{end_row}')
            ws.merge_cells(f'G{start_row}:G{end_row}')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'A{start_row}'].font = Font(name='Arial', size=9, bold=True)
        tc = ws.cell(row=start_row, column=7)
        tc.value = f'=SUM(F{start_row}:F{end_row})'
        tc.number_format = '#,##0'
        tc.alignment = Alignment(horizontal='right', vertical='center')
        tc.font = Font(name='Arial', size=9, bold=True)
    total_row = row
    ws.merge_cells(f'A{total_row}:E{total_row}')
    ws[f'A{total_row}'] = '合　計'
    ws[f'A{total_row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'A{total_row}'].fill = PatternFill('solid', fgColor=TOTAL_BG)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{total_row}'].border = medium_border()
    for col in (6, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f'=SUM(F3:F{total_row-1})'
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = PatternFill('solid', fgColor=TOTAL_BG)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = medium_border()
    ws.row_dimensions[total_row].height = 24
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 22
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.freeze_panes = 'A3'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, diff

# ==================== フォルダID永続化（Secretsから初期値読み込み）====================
try:
    default_inv = st.secrets["FOLDER_ID_INV"]
except:
    default_inv = ""
try:
    default_csv = st.secrets["FOLDER_ID_CSV"]
except:
    default_csv = ""

# ==================== UI ====================
st.markdown("""
<div class="title-bar">
    <div class="main-title">📊 Indeed請求明細ジェネレーター</div>
    <div class="sub-title">Google DriveのデータからクライアントごとのIndeed請求明細Excelを自動生成します</div>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.header("⚙️ Google Drive設定")
    st.markdown("---")
    folder_id_inv = st.text_input("📁 請求データ（Excel）フォルダID", value=default_inv, placeholder="Google DriveのフォルダIDを入力")
    folder_id_csv = st.text_input("📁 キャンペーンパフォーマンス（CSV）フォルダID", value=default_csv, placeholder="Google DriveのフォルダIDを入力")
    st.markdown("---")
    st.caption("※ サービスアカウントのメールアドレスを各フォルダに共有してください")

month_options = {
    "2026年1月": ("2026-01-01", "Indeed_2026年1月.xlsx"),
    "2026年2月": ("2026-02-01", "Indeed_2026年2月.xlsx"),
    "2026年3月": ("2026-03-01", "Indeed_2026年3月.xlsx"),
}

col1, col2 = st.columns([1, 1], gap="large")

with col1:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("① 対象月を選択")
    selected_months = st.multiselect("対象月（複数選択可）", list(month_options.keys()), default=["2026年3月"])
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("② クライアント名を入力")
    client_name = st.text_input("クライアント名（部分一致）", placeholder="例：JSS、ノムラメディアス、ORES など")
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("③ Driveからデータを取得して生成")

    if st.button("📥 請求明細Excelを生成", use_container_width=True):
        if not folder_id_inv or not folder_id_csv:
            st.error("サイドバーでGoogle DriveのフォルダIDを設定してください")
        elif not client_name:
            st.error("クライアント名を入力してください")
        elif not selected_months:
            st.error("対象月を1つ以上選択してください")
        else:
            with st.spinner("Google Driveからデータを取得中..."):
                try:
                    service = get_drive_service()

                    # CSVを一括取得
                    csv_files = list_files_in_folder(service, folder_id_csv)
                    all_csv_df = None
                    for f in csv_files:
                        if f['name'].endswith('.csv'):
                            try:
                                buf = download_file(service, f['id'])
                                try:
                                    df = pd.read_csv(buf, encoding='utf-8-sig')
                                except:
                                    buf.seek(0)
                                    try:
                                        df = pd.read_csv(buf, encoding='utf-8')
                                    except:
                                        buf.seek(0)
                                        df = pd.read_csv(buf, encoding='utf-16', sep='\t')
                                if '対象年月' in df.columns:
                                    all_csv_df = df if all_csv_df is None else pd.concat([all_csv_df, df], ignore_index=True)
                            except:
                                pass

                    if all_csv_df is None:
                        st.error("キャンペーンパフォーマンスデータが見つかりません")
                        st.stop()

                    all_csv_df = all_csv_df.drop_duplicates()

                    # 月ごとにExcel生成
                    results = []
                    for month_label in selected_months:
                        month_key, inv_filename = month_options[month_label]
                        inv_files = list_files_in_folder(service, folder_id_inv)
                        inv_file = next((f for f in inv_files if f['name'] == inv_filename), None)
                        if not inv_file:
                            st.warning(f"⚠️ {inv_filename} がGoogle Driveに見つかりません（スキップ）")
                            continue
                        inv_buf = download_file(service, inv_file['id'])
                        inv_df = pd.read_excel(inv_buf)
                        csv_month = all_csv_df[all_csv_df['対象年月'] == month_key]
                        if len(csv_month) == 0:
                            st.warning(f"⚠️ {month_label}のキャンペーンデータが見つかりません（スキップ）")
                            continue
                        result_buf, diff = create_billing_excel(client_name, inv_df, csv_month, month_label)
                        if result_buf is None:
                            st.error(diff)
                        else:
                            results.append((month_label, result_buf, diff))

                    if results:
                        st.success(f"データ取得完了！{len(results)}件のExcelを生成しました")
                        for month_label, result_buf, diff in results:
                            fname = f"{client_name}_Indeed請求明細_{month_label}.xlsx"
                            st.download_button(
                                label=f"⬇️ {month_label}　Excelをダウンロード",
                                data=result_buf,
                                file_name=fname,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"dl_{month_label}"
                            )
                            if diff == 0:
                                st.markdown(f'<div class="result-box">✅ {month_label}　突合完了・差異ゼロ</div>', unsafe_allow_html=True)
                            else:
                                st.markdown(f'<div class="error-box">⚠️ {month_label}　差異あり：¥{abs(int(diff)):,}</div>', unsafe_allow_html=True)

                except Exception as e:
                    st.error(f"エラーが発生しました：{e}")

    st.markdown('</div>', unsafe_allow_html=True)

st.markdown("---")
st.caption("※ Google Driveのフォルダにサービスアカウントのメールを「閲覧者」として共有してください")
